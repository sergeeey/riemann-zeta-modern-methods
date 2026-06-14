# -*- coding: utf-8 -*-
"""
Генерация русского Excel-зеркала дорожной карты H-12 (Гипотеза Римана).
Источник истины — ROADMAP.md (в git). Этот скрипт зеркалит его в xlsx для личного просмотра.
WHY: xlsx бинарный и живёт вне git → данные держим здесь, при обновлении прогресса
правим ROADMAP.md И этот блок данных, затем перезапускаем скрипт.

Запуск:  python scripts/build_roadmap_xlsx.py
Выход:   C:\\Users\\serge\\Downloads\\H12_Riemann_Roadmap_RU.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = r"C:\Users\serge\Downloads\H12_Riemann_Roadmap_RU.xlsx"

# ---- палитра ----
DARK = PatternFill("solid", fgColor="1F3864")
MID = PatternFill("solid", fgColor="2E5395")
BLOCK = PatternFill("solid", fgColor="D6DCE4")
GREEN = PatternFill("solid", fgColor="C6EFCE")
BLUE = PatternFill("solid", fgColor="BDD7EE")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
GREY = PatternFill("solid", fgColor="E7E6E6")
RED = PatternFill("solid", fgColor="F8CBAD")
CMAP = {"green": GREEN, "blue": BLUE, "yellow": YELLOW, "grey": GREY, "red": RED}

WHITE_BOLD = Font(color="FFFFFF", bold=True, size=11)
WHITE_BIG = Font(color="FFFFFF", bold=True, size=13)
BOLD = Font(bold=True, size=10)
NORM = Font(size=10)
ITAL = Font(color="FFFFFF", size=9, italic=True)
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAPC = Alignment(wrap_text=True, vertical="center", horizontal="center")

# ---- данные (зеркало ROADMAP.md, RU) ----
TITLE = (
    "H-12 · Гипотеза Римана — Дорожная карта и журнал прогресса   |   "
    "Обновлено: 2026-06-14   |   Методология: FL Full-Ladder + EstimandOps L0 + skeptic"
)
CONSTRAINTS = (
    "ЖЁСТКИЕ ОГРАНИЧЕНИЯ:  фундамент = классика (1859 / 1896 / 1901), не пересматриваем  ·  "
    "Grant papers ≠ фундамент (REJECTED, см. null_results)  ·  "
    "любое «100% / доказано / точно» → запуск skeptic  ·  статус мира ≠ наша проработка"
)

ATOM_HDR = ["#", "Атом", "Мир", "Наша %", "Артефакт", "Что даёт RH · что НЕ значит"]
ATOMS = [
    (
        "A-01",
        "Функциональное уравнение ζ(s)=2^s·π^{s−1}·sin(πs/2)·Γ(1−s)·ζ(1−s)",
        "✅ Риман 1859",
        "100% (фундамент)",
        "foundation §3",
        "Даёт симметрию относит. Re=½ (ρ→1−ρ). НЕ доказывает, что нули на ½ — объясняет «почему ½».",
        "green",
    ),
    (
        "A-02",
        "Тривиальные нули s=−2,−4,−6,…",
        "✅",
        "100% (фундамент)",
        "foundation §4",
        "Закрытый вопрос, базовый контроль верификации. НЕ относится к нетривиальным нулям.",
        "green",
    ),
    (
        "A-03",
        "Критическая полоса 0<Re<1",
        "✅ Адамар/ВП 1896",
        "100% (фундамент)",
        "foundation §5",
        "Сужает зону нулей; даёт ТПЧ. НЕ доказывает Re=½ (½ лишь внутри полосы).",
        "green",
    ),
    (
        "A-04",
        "Бесконечно много нулей на Re=½ (функция Харди Z(t))",
        "✅ Харди 1914",
        "0%",
        "—",
        "«Некоторые лежат на ½». НЕ = «все лежат» — здесь главный gap.",
        "green",
    ),
    (
        "A-05",
        "Связь с π(x): явная формула + von Koch",
        "✅ 1901",
        "0%",
        "—",
        "RH ⟺ |π(x)−Li(x)|=O(√x·log x). Эквивалентность — осторожно: эту оценку нельзя постулировать (см. grant-audit).",
        "green",
    ),
    (
        "A-06",
        "Эквивалентные формулировки (Robin, Li, Mertens)",
        "✅",
        "0%",
        "—",
        "Альтернативные пути атаки. НЕ известно, что какой-то легче исходного.",
        "green",
    ),
    (
        "A-07",
        "Доля нулей на Re=½ (>41.28%)",
        "🟡 Прэт 2013",
        "0%",
        "—",
        "Активная область, каждый % = прорыв. Gap 41%→100% без мостика.",
        "yellow",
    ),
    (
        "A-08",
        "Полоса, свободная от нулей, у Re=1",
        "🟡",
        "0%",
        "—",
        "Расширить до Re>½+ε ⟹ доказать RH. Сейчас полоса узкая.",
        "yellow",
    ),
    (
        "A-09",
        "RMT-статистика нулей (GUE, Монтгомери–Одлыко)",
        "🟡",
        "0%",
        "—",
        "База подхода Гильберта–Пойи. Совпадение наблюдается, НЕ доказано.",
        "yellow",
    ),
    (
        "A-10",
        "Вычислительная верификация (первые 2×10¹³ нулей)",
        "🧮",
        "100% ✅ ЗАВЕРШЁН (PROMOTE)",
        "experiments/20260614-A10-computational-base/",
        "Верифиц. база + toolkit src/zeta. НЕ доказательство (конечная проверка).",
        "blue",
    ),
    (
        "A-11",
        "САМА ГИПОТЕЗА: ∀ρ нетрив. → Re=½",
        "❓",
        "0%",
        "—",
        "Финальная цель = синтез A-07/A-08/A-12. Прямая атака — последней.",
        "grey",
    ),
    (
        "A-12",
        "Оператор Гильберта–Пойи (спектральный подход)",
        "❓",
        "0%",
        "—",
        "Спектр самосопряж. H веществен ⟹ Re=½. ВЫСОКИЙ приоритет. Мост к S³-методологии.",
        "grey",
    ),
    (
        "A-13",
        "Обобщённая RH (L-функции, GRH)",
        "❓",
        "0%",
        "—",
        "Сложнее RH. Не фокус сейчас.",
        "grey",
    ),
    (
        "A-14",
        "Геометрический/физический подход",
        "❓ [HYPOTHESIS]",
        "Grant под-часть REJECTED",
        "null_results/20260614-grant-rh-audit.md",
        "Grant (Eisenstein/iHarmonic) забракован. Остаются: квант. хаос, билиарды, Selberg. Phase 4.",
        "red",
    ),
]

INS_HDR = ["#", "Наблюдение", "Статус", "Ядро / Evidence"]
INSIGHTS = [
    (
        "1",
        "src/zeta toolkit построен и независимо валидирован (skeptic + reviewer)",
        "✅",
        "100 нулей на Re=½, Turing-полнота до T≈237, off-line детектор до T≈66. experiments/20260614-A10",
    ),
    (
        "2",
        "Оба Grant-доказательства RH не проходят",
        "✅ REJECT",
        "Doc A циркулярен (Step 4 = von Koch = сама RH); Doc B = недостроенный Гильберт–Пойя (Spectral Isomorphism = Conjecture). Thm 5.3 численно ЛОЖНА [VERIFIED-tool]",
    ),
    (
        "3",
        "S³-Spinor журнал = образец честной работы с оператором → шаблон для A-12",
        "🟡 candidate",
        "Спектр верифицируется (E0 gate), negative controls (NC-2), λ помечен FREE, overclaim BLOCKED. Браться за A-12 — вести так же.",
    ),
]

Q_HDR = ["Q#", "Вопрос", "Зачем", "Что даст ответ"]
QUESTIONS = [
    (
        "Q1",
        "Где ceiling метода Левинсона/Конри (A-07)? Почему ~41%?",
        "Понять, исчерпан ли метод",
        "Есть запас → атакуемый %; ceiling → нужен другой класс методов",
    ),
    (
        "Q2",
        "Какие кандидаты на оператор H существуют — Berry–Keating, Connes, de Branges (A-12)?",
        "Wheels First — не строить с нуля",
        "Список операторов + можно ли верифицировать спектр строго (как S³)",
    ),
    (
        "Q3",
        "Методы расширения zero-free region и их предел (A-08)?",
        "Прямой путь к RH через ширину полосы",
        "Понять, упирается ли в известный барьер",
    ),
    (
        "Q4",
        "Lit-search: свежие (<2 лет) результаты по A-07/A-09/A-12",
        "Не повторять сделанное",
        "Карта фронта; куда приложить вычислит. базу A-10",
    ),
]

# ---- рендеринг ----
wb = Workbook()
ws = wb.active
ws.title = "Журнал H-12"
for i, w in enumerate([6, 40, 18, 22, 34, 60], 1):
    ws.column_dimensions[chr(64 + i)].width = w

r = 1


def banner(text, fill, font, height):
    """Один объединённый баннер на всю ширину (6 колонок)."""
    global r
    # значение пишем ДО merge, иначе ячейка становится read-only
    ws.cell(r, 1, text)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws.cell(r, 1)
    c.fill, c.font, c.alignment = fill, font, WRAP
    ws.row_dimensions[r].height = height
    r += 1


def row(vals, *, font, fill, align, height, merge_tail_from=None):
    """Строка из 6 колонок. merge_tail_from=N → колонки N..6 объединяются.
    Значения и стили выставляются ДО merge (обходит read-only MergedCell)."""
    global r
    for j in range(1, 7):
        v = (
            vals[j - 1]
            if (j - 1) < len(vals) and (merge_tail_from is None or j <= merge_tail_from)
            else ""
        )
        c = ws.cell(r, j, v)
        c.font, c.alignment, c.border = font, align, BORDER
        if fill:
            c.fill = fill
    if merge_tail_from:
        ws.merge_cells(start_row=r, start_column=merge_tail_from, end_row=r, end_column=6)
    ws.row_dimensions[r].height = height
    r += 1


banner(TITLE, DARK, WHITE_BIG, 42)
banner(CONSTRAINTS, MID, ITAL, 44)
r += 1

banner("БЛОК 1 — Журнал прогресса по атомам (A-01 … A-14)", BLOCK, BOLD, 20)
row(ATOM_HDR, font=WHITE_BOLD, fill=MID, align=WRAPC, height=26)
for *vals, color in ATOMS:
    row(vals, font=NORM, fill=CMAP[color], align=WRAP, height=72)
r += 1

banner("БЛОК 2 — Ключевые инсайты (кросс-сессионная дистилляция)", BLOCK, BOLD, 20)
row(INS_HDR, font=WHITE_BOLD, fill=MID, align=WRAPC, height=26, merge_tail_from=4)
for rec in INSIGHTS:
    row(list(rec), font=NORM, fill=None, align=WRAP, height=74, merge_tail_from=4)
r += 1

banner("БЛОК 3 — Открытые вопросы (математические мишени)", BLOCK, BOLD, 20)
row(Q_HDR, font=WHITE_BOLD, fill=MID, align=WRAPC, height=26, merge_tail_from=4)
for rec in QUESTIONS:
    row(list(rec), font=NORM, fill=None, align=WRAP, height=62, merge_tail_from=4)

ws.freeze_panes = "A6"
wb.save(OUT)
print("OK ->", OUT)
print("rows:", r - 1)
